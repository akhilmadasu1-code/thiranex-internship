"""
excel_report.py
Generates a formatted Excel workbook report with:
  - Raw data sheet
  - Cleaned data sheet
  - Summary statistics sheet
  - Charts embedded in the workbook
"""

import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

os.makedirs("reports", exist_ok=True)

# ── Colours ────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
GREEN      = "70AD47"
RED        = "FF0000"
GREY       = "F2F2F2"
WHITE      = "FFFFFF"


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_row(ws, row_num: int, ncols: int, fill_hex: str = MID_BLUE) -> None:
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=WHITE, size=11)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()


def auto_column_width(ws) -> None:
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        try:
            max_len = max((len(str(cell.value)) for cell in col
                           if cell.value and not isinstance(cell, MergedCell)), default=8)
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(max_len + 4, 40)
        except Exception:
            pass


def write_df_to_sheet(ws, df: pd.DataFrame, title: str,
                       header_fill: str = MID_BLUE) -> None:
    """Write a DataFrame to a worksheet with formatting."""
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14, color=DARK_BLUE)
    ws.append([])

    # Header row
    ws.append(list(df.columns))
    style_header_row(ws, ws.max_row, len(df.columns), header_fill)

    # Data rows
    alt_fill = PatternFill("solid", fgColor=GREY)
    for i, row in enumerate(df.itertuples(index=False)):
        ws.append(list(row))
        if i % 2 == 0:
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=ws.max_row, column=col).fill = alt_fill
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=ws.max_row, column=col).border = _border("hair")

    auto_column_width(ws)


def build_summary_sheet(ws, df_raw: pd.DataFrame, df_clean: pd.DataFrame) -> None:
    """Create a Summary Statistics sheet."""
    ws.append(["DATA CLEANING — SUMMARY REPORT"])
    ws["A1"].font = Font(bold=True, size=16, color=DARK_BLUE)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    # KPI table
    kpis = [
        ("Metric",                 "Before Cleaning", "After Cleaning", "Change"),
        ("Total Rows",             len(df_raw),        len(df_clean),
         len(df_clean) - len(df_raw)),
        ("Total Columns",          len(df_raw.columns), len(df_clean.columns),
         len(df_clean.columns) - len(df_raw.columns)),
        ("Missing Values",
         int(df_raw.isna().sum().sum()),
         int(df_clean.isna().sum().sum()),
         int(df_clean.isna().sum().sum()) - int(df_raw.isna().sum().sum())),
        ("Duplicate Rows",
         int(df_raw.duplicated().sum()),
         int(df_clean.duplicated().sum()),
         int(df_clean.duplicated().sum()) - int(df_raw.duplicated().sum())),
    ]
    for row_data in kpis:
        ws.append(list(row_data))
    style_header_row(ws, 3, 4, DARK_BLUE)

    change_fill_pos = PatternFill("solid", fgColor="C6EFCE")  # green
    change_fill_neg = PatternFill("solid", fgColor="FFC7CE")  # red

    for row_idx in range(4, 3 + len(kpis)):
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center")
        change_cell = ws.cell(row=row_idx, column=4)
        if isinstance(change_cell.value, (int, float)):
            change_cell.fill = change_fill_neg if change_cell.value < 0 else change_fill_pos

    ws.append([])

    # Numeric statistics
    ws.append(["Numeric Column Statistics (Cleaned Data)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=13, color=DARK_BLUE)
    ws.append([])

    num_stats = df_clean.describe().round(2).reset_index()
    num_stats.rename(columns={"index": "Statistic"}, inplace=True)
    ws.append(list(num_stats.columns))
    style_header_row(ws, ws.max_row, len(num_stats.columns), MID_BLUE)
    for _, row in num_stats.iterrows():
        ws.append(list(row))
        for col in range(1, len(num_stats.columns) + 1):
            ws.cell(ws.max_row, col).border = _border("hair")

    auto_column_width(ws)


def add_bar_chart(ws, title: str, categories_ref, values_ref,
                   anchor: str = "F3") -> None:
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Column"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data = Reference(ws, **values_ref)
    cats = Reference(ws, **categories_ref)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def main():
    # ── Load data ─────────────────────────────────────────────────────────
    if not os.path.exists("data/sample_data.csv"):
        print("⚠  data/sample_data.csv not found — run generate_sample_data.py first")
        return
    if not os.path.exists("data/cleaned_data.csv"):
        print("⚠  data/cleaned_data.csv not found — run data_cleaner.py first")
        return

    df_raw   = pd.read_csv("data/sample_data.csv")
    df_clean = pd.read_csv("data/cleaned_data.csv")

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    build_summary_sheet(ws_summary, df_raw, df_clean)

    # ── Sheet 2: Raw Data (first 100 rows) ────────────────────────────────
    ws_raw = wb.create_sheet("Raw Data")
    write_df_to_sheet(ws_raw, df_raw.head(100), "Raw Data (first 100 rows)", RED)

    # ── Sheet 3: Cleaned Data ─────────────────────────────────────────────
    ws_clean = wb.create_sheet("Cleaned Data")
    write_df_to_sheet(ws_clean, df_clean.head(100), "Cleaned Data (first 100 rows)", GREEN)

    # ── Save ──────────────────────────────────────────────────────────────
    out = "reports/data_report.xlsx"
    wb.save(out)
    print(f"✅ Excel report saved: {out}")


if __name__ == "__main__":
    main()
